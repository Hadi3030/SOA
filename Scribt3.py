import streamlit as st
import pandas as pd
import io
import re

def parse_ref(ref):
    match = re.match(r"(\d+)(/.+)", ref)
    if match:
        return int(match.group(1)), match.group(2)
    return 1, ""

def build_ref(num, suffix):
    return f"{num}{suffix}"
    
st.set_page_config(page_title="SOA Report ALL Broker", layout="wide")

# LOGO + TITLE
col1, col2 = st.columns([1, 6])

with col1:
    st.image("askrindo.jpg", width=120)

with col2:
    st.title("📑 SOA Report Generator - All Broker")

# ===============================
# UPLOAD
# ===============================
file = st.file_uploader("Upload File SOA", type=["xlsx","xlsb"])
if not file:
    st.stop()

# excel_file = pd.ExcelFile(file)
# sheet = st.selectbox("Pilih Sheet", excel_file.sheet_names)
# df = pd.read_excel(excel_file, sheet_name=sheet)

# DETECT FILE TYPE
if file.name.endswith(".xlsb"):
    excel_file = pd.ExcelFile(file, engine="pyxlsb")
else:
    excel_file = pd.ExcelFile(file)

# PILIH SHEET
sheet = st.selectbox("Pilih Sheet", excel_file.sheet_names)

# BACA DATA
if file.name.endswith(".xlsb"):
    df = pd.read_excel(file, sheet_name=sheet, engine="pyxlsb")
else:
    df = pd.read_excel(file, sheet_name=sheet)

# ===============================
# NORMALISASI KOLOM
# ===============================
df.columns = df.columns.str.strip().str.lower()

mapping = {
    'prod':'PROD','cob':'COB','uy':'UY','curr':'CURRENCY',
    'broker':'BROKER',
    'qs_ceding':'QS_CEDING','sp_ceding':'SP_CEDING',
    'komisi_qs':'KOMISI_QS','komisi_sp':'KOMISI_SP',
    'klaim_qs':'KLAIM_QS','klaim_sp':'KLAIM_SP'
}
df = df.rename(columns=mapping)

# ===============================
# FIX TEXT
# ===============================
for col in ['CURRENCY', 'COB', 'BROKER']:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.upper()

df = df[(df['CURRENCY'] != "") & (df['CURRENCY'].notna())]

# ===============================
# HANDLE UY (*)
# ===============================
if "UY" in df.columns:
    df['UY_ORIGINAL'] = df['UY'].astype(str)
    df['UY_FLAG'] = df['UY_ORIGINAL'].str.contains(r'\*')

    # untuk perhitungan (hapus *)
    df['UY'] = df['UY_ORIGINAL'].str.replace(r'\*', '', regex=True)
    df['UY'] = pd.to_numeric(df['UY'], errors='coerce')

    # untuk tampilan
    df['UY_DISPLAY'] = df['UY_ORIGINAL']
    
# ===============================
# FILTER BROKER
# ===============================
st.subheader("Pilih Broker")

if "BROKER" in df.columns:
    broker_list = sorted(df["BROKER"].dropna().unique().tolist())

    selected_broker = broker_list if st.checkbox("ALL BROKER", True) else [
        b for b in broker_list if st.checkbox(b)
    ]

    df = df[df["BROKER"].isin(selected_broker)]

# ===============================
# FILTER COB
# ===============================
st.subheader("Pilih COB")

if "COB" in df.columns:
    cob_list = sorted(df["COB"].dropna().unique().tolist())

    selected_cob = cob_list if st.checkbox("ALL COB", True) else [
        cob for cob in cob_list if st.checkbox(cob)
    ]

    df = df[df["COB"].isin(selected_cob)]

# ===============================
# FILTER UW YEAR
# ===============================
st.subheader("Pilih UW Year")

if "UY" in df.columns:
    uy_list = sorted(df["UY"].dropna().unique().tolist())

    selected_uy = uy_list if st.checkbox("ALL UW YEAR", True) else [
        uy for uy in uy_list if st.checkbox(str(uy))
    ]

    df = df[df["UY"].isin(selected_uy)]

# ===============================
# FILTER LT
# ===============================
lt_option = st.selectbox("Filter Long Term", ["ALL", "LT", "NON-LT"])

if lt_option != "ALL":
    if lt_option == "LT":
        df = df[df['COB'].str.contains("LT", na=False)]
    else:
        df = df[~df['COB'].str.contains("LT", na=False)]

zero_option = st.selectbox(
    "Tampilkan Baris Nol",
    ["Show All", "Hide Zero Rows"]
)

# ===============================
# CLEAN NUMERIC
# ===============================
def clean_number(x):
    if pd.isna(x):
        return 0

    # kalau sudah numeric → jangan diapa-apain
    if isinstance(x, (int, float)):
        return float(x)

    x = str(x).strip().replace(" ", "")

    # format Indonesia
    if ',' in x and '.' in x:
        if x.rfind(',') > x.rfind('.'):
            x = x.replace('.', '').replace(',', '.')
        else:
            x = x.replace(',', '')
    elif ',' in x:
        x = x.replace(',', '.')
    
    try:
        return float(x)
    except:
        return 0
# def clean_number(x):
#     if pd.isna(x):
#         return 0
    
#     x = str(x).strip()
    
#     # format Indonesia: 1.234.567,89 → 1234567.89
#     x = x.replace('.', '').replace(',', '.')
    
#     try:
#         return float(x)
#     except:
#         return 0
        
num_cols = [
    'QS_CEDING','SP_CEDING',
    'KOMISI_QS','KOMISI_SP',
    'KLAIM_QS','KLAIM_SP'
]

# for col in num_cols:
#     df[col] = (
#         df[col].astype(str)
#         .str.replace('.', '', regex=False)
#         .str.replace(',', '.', regex=False)
#     )
#     df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
for col in num_cols:
    if col in df.columns:
        df[col] = df[col].apply(clean_number)

# ===============================
# PARSE PROD
# ===============================
def parse_prod(x):
    try:
        x = str(x)
        return int(x[:4]), int(x[-2:])
    except:
        return None, None

df[['YEAR','MONTH']] = df['PROD'].apply(lambda x: pd.Series(parse_prod(x)))
df = df.dropna(subset=['YEAR','MONTH'])

# ===============================
# DATE INFO
# ===============================
month_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
             7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

min_m = int(df['MONTH'].min())
max_m = int(df['MONTH'].max())
year = int(df['YEAR'].mode()[0])

months_text = f"{month_map[min_m]} - {month_map[max_m]} {year}"

def get_quarter(m):
    return ["I","II","III","IV"][(m-1)//3]

quarter = get_quarter(max_m)

has_star_year = df['PROD'].astype(str).str.contains(r'\*').any()

# ===============================
# GENERATE REPORT
# ===============================
def generate_report(df, tipe, zero_option):

    if tipe == "QS":
        df['PREMIUM'] = df['QS_CEDING']
        df['COMMISSION'] = df['KOMISI_QS']
        df['CLAIM'] = df['KLAIM_QS']
    else:
        df['PREMIUM'] = df['SP_CEDING']
        df['COMMISSION'] = df['KOMISI_SP']
        df['CLAIM'] = df['KLAIM_SP']

    grouped = (
        df.groupby(['CURRENCY','COB','UY','UY_DISPLAY'])
        .sum(numeric_only=True)
        .reset_index()
        .sort_values(['CURRENCY','COB','UY','UY_DISPLAY'])
    )
    
    grouped['AMOUNT'] = (
        grouped['PREMIUM']
        - grouped['COMMISSION']
        - grouped['CLAIM']
    )

    rows = []

    for curr, df_curr in grouped.groupby('CURRENCY'):

        if zero_option == "Hide Zero Rows":
            df_curr = df_curr[
                ~((df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']] == 0).all(axis=1))
            ]

        if df_curr.empty:
            continue

        first_row = True

        for cob, df_cob in df_curr.groupby('COB'):

            if df_cob.empty:
                continue

            first_cob_row = True

            for _, r in df_cob.iterrows():
                rows.append([
                    curr if first_row else "",
                    cob if first_cob_row else "",
                    r['UY_DISPLAY'],
                    r['PREMIUM'],
                    r['COMMISSION'],
                    r['CLAIM'],
                    r['AMOUNT']
                ])
                first_row = False
                first_cob_row = False

            subtotal = df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
            rows.append(["", f"{cob} TOTAL", "", *subtotal])

        total_curr = df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
        rows.append([f"{curr} TOTAL","","", *total_curr])
        rows.append(["","","","","","",""])

    return pd.DataFrame(
        rows,
        columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','AMOUNT']
    )

# ===============================
# INPUT USER
# ===============================
ref_qs = st.text_input("Ref No QS")
ref_sp = st.text_input("Ref No SPL")
remarks = st.text_area("Remarks")
note = st.text_area("Note")

signature_place = st.text_input("Tempat Tanda Tangan", value="Jakarta, ........")
file_name = st.text_input("Nama file", value="SOA_Report")
# ===============================
# EXPORT MULTI BROKER
# ===============================
if st.button("⬇️ Download All Broker"):

    output = io.BytesIO()

    from openpyxl.styles import Font, Alignment, PatternFill, Border
    from openpyxl.drawing.image import Image

    header_fill = PatternFill("solid", fgColor="000000")
    grey_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    no_border = Border()
    

    def write_sheet(writer, data, name, tipe, ref, broker_name, signature_place):

        data.to_excel(writer, index=False, sheet_name=name, startrow=12)
        ws = writer.sheets[name]
    
        # ======================
        # LOGO
        # ======================
        try:
            logo = Image("askrindo.jpg")
            logo.height = 60
            logo.width = 140
            ws.add_image(logo, "A1")
        except:
            pass
    
        # ======================
        # TITLE
        # ======================
        ws.merge_cells('A4:G4')
        ws['A4'] = "STATEMENT OF ACCOUNT"
        ws['A4'].font = Font(bold=True, size=14)
        ws['A4'].alignment = Alignment(horizontal='center')
    
        ws.merge_cells('A5:G5')
        ws['A5'] = f"Ref No. {ref}"
        ws['A5'].font = Font(bold=True)
        ws['A5'].alignment = Alignment(horizontal='center')
    
        # ======================
        # HEADER INFO
        # ======================
        ws['A7'] = "Treaty Year  :"
        ws['B7'] = year
    
        ws['A8'] = "Quarter      :"
        ws['B8'] = f"{quarter} {tipe}"
    
        ws['A9'] = "Contract Name:"
        contract_name = "Quota Share" if tipe.strip().upper() == "QS" else "Surplus"
        ws['B9'] = contract_name
    
        ws['A10'] = "For Months   :"
        ws['B10'] = months_text

        ws['A11'] = "Remarks      :"
        ws['B11'] = remarks
    
        # ======================
        # TABLE HEADER STYLE
        # ======================
        header_fill = PatternFill("solid", fgColor="000000")
        header_font = Font(color="FFFFFF", bold=True)
    
        for col in range(1, 8):
            cell = ws.cell(row=13, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
        # ======================
        # BODY FORMAT
        # ======================
        current_currency = None
    
        for row in range(14, ws.max_row + 1):
    
            val_curr = ws[f"A{row}"].value
    
            for col in "ABCDEFG":
                ws[f"{col}{row}"].fill = white_fill
    
            if all(ws[f"{col}{row}"].value in ["", None] for col in "ABCDEFG"):
                current_currency = None
                continue
    
            if val_curr and "TOTAL" not in str(val_curr):
                current_currency = val_curr
    
            if current_currency:
                ws[f"A{row}"].fill = grey_fill
    
            if val_curr and "TOTAL" in str(val_curr):
                for col in "ABCDEFG":
                    ws[f"{col}{row}"].fill = grey_fill
                current_currency = None
    
            for col in ['D','E','F','G']:
                ws[f"{col}{row}"].number_format = '#,##0.00;[Red](#,##0.00)'
    
        # ======================
        # NOTE
        # ======================
        last = ws.max_row + 2
        if has_star_year:
            ws[f"A{last}"] = "Note :"
            ws[f"B{last}"] = note
    
        # ======================
        # FOOTER TTD (INI YANG KAMU TANYA)
        # ======================
        footer_row = last + 4
    
        # LEFT: BROKER
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=3)
        ws.cell(row=footer_row, column=1).value = f"{broker_name}"
        ws.cell(row=footer_row, column=1).font = Font(bold=True)
        ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal="left")
    
        # RIGHT: PLACE + SIGNATURE
        ws.merge_cells(start_row=footer_row, start_column=5, end_row=footer_row, end_column=7)
        ws.cell(row=footer_row, column=5).value = signature_place
        ws.cell(row=footer_row, column=5).alignment = Alignment(horizontal="right")
    
        # SIGNATURE LINE
        ws.merge_cells(start_row=footer_row + 2, start_column=5, end_row=footer_row + 2, end_column=7)
        ws.cell(row=footer_row + 2, column=5).value = ".............................."
        ws.cell(row=footer_row + 2, column=5).alignment = Alignment(horizontal="right")

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        brokers = df["BROKER"].dropna().unique()
        qs_start, qs_suffix = parse_ref(ref_qs)
        sp_start, sp_suffix = parse_ref(ref_sp)
        ref_counter = qs_start  # GLOBAL SEQUENCE
        
        qs_counter = qs_start
        sp_counter = sp_start
       
        for broker in brokers:
            df_b = df[df["BROKER"] == broker]
        
            qs_ref = build_ref(ref_counter, qs_suffix)
            ref_counter += 1
        
            sp_ref = build_ref(ref_counter, sp_suffix)
            ref_counter += 1
        
            report_qs = generate_report(df_b.copy(), "QS", zero_option)
            report_sp = generate_report(df_b.copy(), "SP", zero_option)
        
            write_sheet(writer, report_qs, f"QS_{broker}"[:31], "QS", qs_ref, broker, signature_place)
            write_sheet(writer, report_sp, f"SP_{broker}"[:31], "SP", sp_ref, broker, signature_place)
            qs_counter += 1
            sp_counter += 1

    st.download_button(
        "📥 Download Excel",
        data=output.getvalue(),
        file_name=f"{file_name}_ALL.xlsx"
    )
