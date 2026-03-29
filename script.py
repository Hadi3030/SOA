import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="SOA Processing", layout="wide")

# ===============================
# SIDEBAR MENU
# ===============================
st.sidebar.title("📌 Menu")

mode = st.sidebar.radio(
    "Pilih Analisis",
    ["Spreading Data (SOA Processing)", "Laporan SOA (SOA Report)"]
)

# ===============================
# SESSION STATE
# ===============================
if "result_data" not in st.session_state:
    st.session_state["result_data"] = None

# ===============================
# FUNCTION PROCESSING
# ===============================
def process_data(df1, df2):

    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    df1['QS'] = pd.to_numeric(df1['QS'], errors='coerce').fillna(0)
    df1['SPL'] = pd.to_numeric(df1['SPL'], errors='coerce').fillna(0)
    df1 = df1[~((df1['QS'] == 0) & (df1['SPL'] == 0))]
    df1['COB'] = df1['COB'].astype(str).str.strip().str.upper()

    cols_convert = ['TSI SHARE','OR','QS','SPL']
    for col in cols_convert:
        df1[col] = pd.to_numeric(df1[col], errors='coerce')

    df1['UY-COB'] = df1['UY'].astype(str) + "-" + df1['COB']

    df2.columns = df2.columns.str.lower()
    df2['uy'] = df2['uy'].astype(str).str.strip()

    for col in ['broker','cob','group']:
        df2[col] = df2[col].astype(str).str.strip().str.upper()

    def percent_to_decimal(x):
        if pd.isna(x):
            return 0
        if isinstance(x, str):
            x = x.replace('%','').strip()
            val = pd.to_numeric(x, errors='coerce')
            return val / 100 if val is not None else 0
        if isinstance(x, (int,float)):
            if x > 1:
                return x / 100
        return x

    df2['sharere']  = df2['sharere'].apply(percent_to_decimal)
    df2['komisiqs'] = df2['komisiqs'].apply(percent_to_decimal)
    df2['komisisp'] = df2['komisisp'].apply(percent_to_decimal)

    df2['cashcall'] = df2['cashcall'].astype(str).str.replace(',', '')
    df2['cashcall'] = pd.to_numeric(df2['cashcall'], errors='coerce')

    merged = df1.merge(
        df2,
        left_on=['UY','COB'],
        right_on=['uy','cob'],
        how='left'
    )

    # SPREAD 2023
    found = merged[merged['broker'].notna()].copy()
    missing = merged[merged['broker'].isna()].copy()

    df2_2023 = df2[df2['uy'] == '2023'].copy()

    missing_expanded = missing.drop(columns=df2.columns, errors='ignore').merge(
        df2_2023,
        left_on='COB',
        right_on='cob',
        how='left'
    )

    merged = pd.concat([found, missing_expanded], ignore_index=True)

    merged = merged.drop(columns=['cashcall','cob','uy'], errors='ignore')

    cols = ['QS','SPL','sharere','komisiqs','komisisp']
    for c in cols:
        merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0)

    merged['qs_ceding'] = merged['QS'] * merged['sharere']
    merged['komisi_qs'] = merged['qs_ceding'] * merged['komisiqs']

    merged['sp_ceding'] = merged['SPL'] * merged['sharere']
    merged['komisi_sp'] = merged['sp_ceding'] * merged['komisisp']

    # ✅ FIX KOLOM (PENTING)
    merged.columns = merged.columns.str.strip().str.upper()

    return merged

# ===============================
# FUNCTION TOTAL
# ===============================
def add_total_row(df):
    numeric_cols = df.select_dtypes(include='number').columns
    total = df[numeric_cols].sum()
    total_df = pd.DataFrame([total])
    total_df.index = ['TOTAL']
    return total_df

# ===============================
# FUNCTION REPORT
# ===============================
def generate_report(df):

    # ✅ FIX KOLOM (PENTING)
    df.columns = df.columns.str.strip().str.upper()

    # ✅ VALIDASI
    required_cols = ['CURRENCY','COB','UY','QS','SPL','KOMISI_QS','KOMISI_SP']
    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        st.error(f"Kolom tidak ditemukan: {missing}")
        st.stop()

    df['PREMIUM'] = df['QS'] + df['SPL']
    df['COMMISSION'] = df['KOMISI_QS'] + df['KOMISI_SP']
    df['CLAIM'] = 0
    df['AMOUNT'] = df['PREMIUM'] - df['COMMISSION']

    grouped = df.groupby(['CURRENCY','COB','UY']).agg({
        'PREMIUM':'sum',
        'COMMISSION':'sum',
        'CLAIM':'sum',
        'AMOUNT':'sum'
    }).reset_index()

    final_rows = []

    for curr in grouped['CURRENCY'].unique():
        df_curr = grouped[grouped['CURRENCY'] == curr]

        for cob in df_curr['COB'].unique():
            df_cob = df_curr[df_curr['COB'] == cob]

            for _, row in df_cob.iterrows():
                final_rows.append([
                    curr, cob, row['UY'],
                    row['PREMIUM'], row['COMMISSION'],
                    row['CLAIM'], row['AMOUNT']
                ])

            total = df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
            final_rows.append(["", f"{cob} TOTAL", "", *total])

        total_curr = df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
        final_rows.append([f"{curr} TOTAL","","", *total_curr])

    return pd.DataFrame(final_rows,
        columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','AMOUNT'])

# ===============================
# MODE 1
# ===============================
if mode == "Spreading Data (SOA Processing)":

    st.title("📊 SOA Processing")

    file_soa = st.file_uploader("Upload Data SOA", type=["xlsx"])
    file_sor = st.file_uploader("Upload SOR Summary", type=["xlsx"])

    if file_soa and file_sor:

        df1 = pd.read_excel(file_soa)
        df2 = pd.read_excel(file_sor)

        st.subheader("Data SOA")
        st.dataframe(df1)

        result = process_data(df1.copy(), df2.copy())

        st.subheader("Hasil Processing")
        st.dataframe(result)

        st.session_state["result_data"] = result

        st.subheader("Total SOA")
        st.dataframe(add_total_row(df1))

        st.subheader("Total Output")
        st.dataframe(add_total_row(result))

        file_name_input = st.text_input("Nama file", value="SOA_Result")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            result.to_excel(writer, index=False)

        final_filename = file_name_input.strip() or "SOA_Result"

        st.download_button(
            "⬇️ Download",
            data=output.getvalue(),
            file_name=f"{final_filename}.xlsx"
        )

# ===============================
# MODE 2
# ===============================
elif mode == "Laporan SOA (SOA Report)":

    st.title("📑 SOA Report")

    source = st.radio(
        "Sumber Data",
        ["Gunakan hasil sebelumnya", "Upload ulang"]
    )

    if source == "Gunakan hasil sebelumnya":

        if st.session_state["result_data"] is None:
            st.warning("Belum ada data, proses dulu")
            st.stop()

        df = st.session_state["result_data"]

    else:
        file = st.file_uploader("Upload hasil", type=["xlsx"])
        if file:
            df = pd.read_excel(file)
        else:
            st.stop()

    st.dataframe(df)

    report = generate_report(df)

    st.subheader("Report SOA")
    st.dataframe(report)

    file_name_input = st.text_input("Nama file report", value="SOA_Report")

    output = io.BytesIO()
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report.to_excel(writer, index=False, sheet_name='SOA Report')
    
        workbook = writer.book
        worksheet = writer.sheets['SOA Report']
    
        # ===============================
        # FORMAT HEADER
        # ===============================
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
    
        # ===============================
        # FORMAT KOLOM ANGKA
        # ===============================
        number_format = '#,##0.00;[Red](#,##0.00)'
    
        cols_to_format = ['D','E','F','G']  # Premium, Commission, Claim, Amount
    
        for col in cols_to_format:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col}{row}"]
                cell.number_format = number_format
    
        # ===============================
        # AUTO WIDTH
        # ===============================
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
    
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
    
            worksheet.column_dimensions[col_letter].width = max_length + 2

    final_filename = file_name_input.strip() or "SOA_Report"

    st.download_button(
        "⬇️ Download Report",
        data=output.getvalue(),
        file_name=f"{final_filename}.xlsx"
    )
