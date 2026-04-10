import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="SOA Processing", layout="wide")

# ===============================
# SIDEBAR MENU
# ===============================
st.sidebar.title("📌 Menu")

mode = st.sidebar.radio(
    "Pilih Analisis",
    ["Spreading Data (SOA Processing)", "Laporan SOA (SOA Report)"]
)

# ===============================
# SESSION STATE
# ===============================
if "result_data" not in st.session_state:
    st.session_state["result_data"] = None

# ===============================
# FUNCTION PROCESSING
# ===============================
def process_data(df1, df2):

    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    df1['QS'] = pd.to_numeric(df1['QS'], errors='coerce').fillna(0)
    df1['SPL'] = pd.to_numeric(df1['SPL'], errors='coerce').fillna(0)
    df1 = df1[~((df1['QS'] == 0) & (df1['SPL'] == 0))]
    df1['COB'] = df1['COB'].astype(str).str.strip().str.upper()

    cols_convert = ['TSI SHARE','OR','QS','SPL']
    for col in cols_convert:
        df1[col] = pd.to_numeric(df1[col], errors='coerce')

    df1['UY-COB'] = df1['UY'].astype(str) + "-" + df1['COB']

    df2.columns = df2.columns.str.lower()
    df2['uy'] = df2['uy'].astype(str).str.strip()

    for col in ['broker','cob','group']:
        df2[col] = df2[col].astype(str).str.strip().str.upper()

    def percent_to_decimal(x):
        if pd.isna(x):
            return 0
        if isinstance(x, str):
            x = x.replace('%','').strip()
            val = pd.to_numeric(x, errors='coerce')
            return val / 100 if val is not None else 0
        if isinstance(x, (int,float)):
            if x > 1:
                return x / 100
        return x

    df2['sharere']  = df2['sharere'].apply(percent_to_decimal)
    df2['komisiqs'] = df2['komisiqs'].apply(percent_to_decimal)
    df2['komisisp'] = df2['komisisp'].apply(percent_to_decimal)

    # MERGE
    merged = df1.merge(
        df2,
        left_on=['UY','COB'],
        right_on=['uy','cob'],
        how='left'
    )

    # SPREAD 2023
    found = merged[merged['broker'].notna()].copy()
    missing = merged[merged['broker'].isna()].copy()

    df2_2023 = df2[df2['uy'] == '2023'].copy()

    missing_expanded = missing.drop(columns=df2.columns, errors='ignore').merge(
        df2_2023,
        left_on='COB',
        right_on='cob',
        how='left'
    )

    merged = pd.concat([found, missing_expanded], ignore_index=True)

    merged = merged.drop(columns=['cob','uy'], errors='ignore')

    cols = ['QS','SPL','sharere','komisiqs','komisisp']
    for c in cols:
        merged[c] = pd.to_numeric(merged[c], errors='coerce').fillna(0)

    merged['qs_ceding'] = merged['QS'] * merged['sharere']
    merged['komisi_qs'] = merged['qs_ceding'] * merged['komisiqs']

    merged['sp_ceding'] = merged['SPL'] * merged['sharere']
    merged['komisi_sp'] = merged['sp_ceding'] * merged['komisisp']

    # ✅ FIX DUPLICATE COLUMN (PENTING BANGET)
    merged = merged.loc[:, ~merged.columns.duplicated()]

    # ✅ RAPIIKAN KOLOM
    merged.columns = merged.columns.str.strip().str.upper()

    return merged

# ===============================
# FUNCTION TOTAL
# ===============================
def add_total_row(df):
    numeric_cols = df.select_dtypes(include='number').columns
    total = df[numeric_cols].sum()
    total_df = pd.DataFrame([total])
    total_df.index = ['TOTAL']
    return total_df

# ===============================
# FUNCTION REPORT
# ===============================
def generate_report(df):

    df.columns = df.columns.str.strip().str.upper()
    df_qs = df.copy()
    df_qs['TYPE'] = 'QS'
    df_qs['PREMIUM'] = df.get('PREMI_PANEL_QS', 0)
    df_qs['COMMISSION'] = df.get('KOMISI_PANEL_QS', 0)
    
    df_sp = df.copy()
    df_sp['TYPE'] = 'SP'
    df_sp['PREMIUM'] = df.get('PREMI_PANEL_SP', 0)
    df_sp['COMMISSION'] = df.get('KOMISI_PANEL_SP', 0)
    
    df_all = pd.concat([df_qs, df_sp])

    # Mapping PRODUCT → COB
    if 'PRODUCT' in df.columns:
        df['COB'] = df['PRODUCT']

    required_cols = ['CURRENCY','COB','UY','QS','SPL','KOMISI_QS','KOMISI_SP']
    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        st.error(f"Kolom tidak ditemukan: {missing}")
        st.stop()

    df['PREMIUM'] = (
        df.get('PREMI_PANEL_QS', 0) +
        df.get('PREMI_PANEL_SP', 0)
    )
    
    df['COMMISSION'] = (
        df.get('KOMISI_PANEL_QS', 0) +
        df.get('KOMISI_PANEL_SP', 0)
    )
    
    df['CLAIM'] = (
        df.get('KLAIM_PANEL_QS', 0) +
        df.get('KLAIM_PANEL_SP', 0)
    )
    
    df['AMOUNT'] = df['PREMIUM'] - df['COMMISSION'] - df['CLAIM']

    grouped = df.groupby(['CURRENCY','COB','UY']).agg({
        'PREMIUM':'sum',
        'COMMISSION':'sum',
        'CLAIM':'sum',
        'AMOUNT':'sum'
    }).reset_index()

    final_rows = []

    for curr in grouped['CURRENCY'].unique():
        df_curr = grouped[grouped['CURRENCY'] == curr]

        for cob in df_curr['COB'].unique():
            df_cob = df_curr[df_curr['COB'] == cob]

            for _, row in df_cob.iterrows():
                final_rows.append([
                    curr, cob, row['UY'],
                    row['PREMIUM'], row['COMMISSION'],
                    row['CLAIM'], row['AMOUNT']
                ])

            total = df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
            final_rows.append(["", f"{cob} TOTAL", "", *total])

        total_curr = df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
        final_rows.append([f"{curr} TOTAL","","", *total_curr])

    return pd.DataFrame(final_rows,
        columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','AMOUNT'])

# ===============================
# MODE 1
# ===============================
if mode == "Spreading Data (SOA Processing)":

    st.title("📊 SOA Processing")

    file_soa = st.file_uploader("Upload Data SOA", type=["xlsx"])
    file_sor = st.file_uploader("Upload SOR Summary", type=["xlsx"])

    if file_soa and file_sor:

        df1 = pd.read_excel(file_soa)
        df2 = pd.read_excel(file_sor)

        st.subheader("Data SOA")
        st.dataframe(df1)

        result = process_data(df1.copy(), df2.copy())

        st.subheader("Hasil Processing")
        st.dataframe(result)

        st.session_state["result_data"] = result

        st.subheader("Total SOA")
        st.dataframe(add_total_row(df1))

        st.subheader("Total Output")
        st.dataframe(add_total_row(result))

        file_name_input = st.text_input("Nama file", value="SOA_Result")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            result.to_excel(writer, index=False)

        final_filename = file_name_input.strip() or "SOA_Result"

        st.download_button(
            "⬇️ Download",
            data=output.getvalue(),
            file_name=f"{final_filename}.xlsx"
        )

# ===============================
# MODE 2
# ===============================
elif mode == "Laporan SOA (SOA Report)":

    st.title("📑 SOA Report")

    file = st.file_uploader("Upload hasil SOA", type=["xlsx"])
    if not file:
        st.stop()

    df = pd.read_excel(file)

    st.dataframe(df)

    report = generate_report(df)

    st.subheader("Report SOA")
    st.dataframe(report)

    st.subheader("📝 Informasi Laporan")
    ref_no = st.text_input("Ref No", value="")
    treaty_year = st.text_input("Treaty Year", value="2026")
    quarter = st.text_input("Quarter", value="Q1")
    months = st.text_input("For Months", value="Jan - Mar")
    remarks = st.text_input("Remarks", value="-")

    file_name_input = st.text_input("Nama file report", value="SOA_Report")

    output = io.BytesIO()

    from openpyxl.styles import Font, Alignment
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
    
        df.columns = df.columns.str.upper()
    
        # ======================
        # MAPPING PRODUCT → COB
        # ======================
        if 'PRODUCT' in df.columns:
            df['COB'] = df['PRODUCT']
    
        # ======================
        # QS
        # ======================
        df_qs = df.copy()
        df_qs['PREMIUM'] = df_qs.get('PREMI_PANEL_QS', 0)
        df_qs['COMMISSION'] = df_qs.get('KOMISI_PANEL_QS', 0)
        df_qs['CLAIM'] = df_qs.get('KLAIM_PANEL_QS', 0)
        df_qs['AMOUNT'] = df_qs['PREMIUM'] - df_qs['COMMISSION'] - df_qs['CLAIM']
    
        report_qs = df_qs.groupby(['CURRENCY','COB','UY']).agg({
            'PREMIUM':'sum',
            'COMMISSION':'sum',
            'CLAIM':'sum',
            'AMOUNT':'sum'
        }).reset_index()
    
        # ======================
        # SP
        # ======================
        df_sp = df.copy()
        df_sp['PREMIUM'] = df_sp.get('PREMI_PANEL_SP', 0)
        df_sp['COMMISSION'] = df_sp.get('KOMISI_PANEL_SP', 0)
        df_sp['CLAIM'] = df_sp.get('KLAIM_PANEL_SP', 0)
        df_sp['AMOUNT'] = df_sp['PREMIUM'] - df_sp['COMMISSION'] - df_sp['CLAIM']
    
        report_sp = df_sp.groupby(['CURRENCY','COB','UY']).agg({
            'PREMIUM':'sum',
            'COMMISSION':'sum',
            'CLAIM':'sum',
            'AMOUNT':'sum'
        }).reset_index()
    
        # ======================
        # TULIS KE EXCEL
        # ======================
        sheet_name = 'SOA Report'
    
        report_qs.to_excel(writer, index=False, sheet_name=sheet_name, startrow=9)
    
        start_sp = len(report_qs) + 15
        report_sp.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_sp)
    
        ws = writer.sheets[sheet_name]
        ws['A9'] = "QUOTA SHARE"
        ws[f"A{start_sp}"] = "SURPLUS"

        try:
            logo = Image("askrindo.jpg")
            logo.width = 120
            logo.height = 60
            ws.add_image(logo, "A1")
        except:
            pass

        ws.merge_cells('A2:G2')
        ws['A2'] = "STATEMENT OF ACCOUNT"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:G3')
        ws['A3'] = f"Ref No. {ref_no}"
        ws['A3'].alignment = Alignment(horizontal='center')

        ws['A5'] = f"Treaty Year : {treaty_year}"
        ws['A6'] = f"Quarter     : {quarter}"
        ws['A7'] = f"For Months  : {months}"
        ws['A8'] = f"Remarks     : {remarks}"

        number_format = '#,##0.00;[Red](#,##0.00)'
        for col in ['D','E','F','G']:
            for row in range(10, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = number_format

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    final_filename = file_name_input.strip() or "SOA_Report"

    st.download_button(
        "⬇️ Download Report",
        data=output.getvalue(),
        file_name=f"{final_filename}.xlsx"
    )
