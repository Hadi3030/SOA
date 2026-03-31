import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="SOA Report", layout="wide")
st.title("📑 SOA Report Generator")

# ===============================
# UPLOAD
# ===============================
file = st.file_uploader("Upload File SOA", type=["xlsx"])
if not file:
    st.stop()

excel_file = pd.ExcelFile(file)
sheet = st.selectbox("Pilih Sheet", excel_file.sheet_names)
df = pd.read_excel(excel_file, sheet_name=sheet)

# ===============================
# NORMALISASI KOLOM
# ===============================
df.columns = df.columns.str.strip().str.lower()

mapping = {
    'prod':'PROD','cob':'COB','uy':'UY','curr':'CURRENCY',
    'broker':'BROKER','qs_ceding':'QS_CEDING','sp_ceding':'SP_CEDING',
    'komisi_qs':'KOMISI_QS','komisi_sp':'KOMISI_SP'
}
df = df.rename(columns=mapping)

# ===============================
# FIX TEXT
# ===============================
for col in ['CURRENCY', 'COB', 'BROKER']:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.upper()

df = df[(df['CURRENCY'] != "") & (df['CURRENCY'].notna())]

# ===============================
# FILTER
# ===============================
broker_list = df['BROKER'].dropna().unique().tolist()
broker_list.insert(0, "ALL")
selected_broker = st.selectbox("Pilih Broker", broker_list)

if selected_broker != "ALL":
    df = df[df['BROKER'] == selected_broker]

lt_option = st.selectbox("Filter Long Term", ["ALL", "LT", "NON-LT"])

if lt_option != "ALL":
    if lt_option == "LT":
        df = df[df['COB'].str.contains("LT", na=False)]
    else:
        df = df[~df['COB'].str.contains("LT", na=False)]

zero_option = st.selectbox(
    "Tampilkan Baris Nol",
    ["Show All", "Hide Zero Rows"]
)

# ===============================
# CLEAN NUMERIC
# ===============================

num_cols = [
    'QS_CEDING','SP_CEDING',
    'KOMISI_QS','KOMISI_SP',
    'KLAIM_QS','KLAIM_SP'
]

for col in num_cols:
    df[col] = (
        df[col].astype(str)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
    )
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# ===============================
# PARSE PROD
# ===============================
def parse_prod(x):
    try:
        x = str(x)
        return int(x[:4]), int(x[-2:])
    except:
        return None, None

df[['YEAR','MONTH']] = df['PROD'].apply(lambda x: pd.Series(parse_prod(x)))
df = df.dropna(subset=['YEAR','MONTH'])

# ===============================
# DATE INFO
# ===============================
month_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
             7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

min_m = int(df['MONTH'].min())
max_m = int(df['MONTH'].max())
year = int(df['YEAR'].mode()[0])

months_text = f"{month_map[min_m]} - {month_map[max_m]} {year}"

def get_quarter(m):
    return ["I","II","III","IV"][(m-1)//3]

quarter = get_quarter(max_m)

# ===============================
# GENERATE REPORT
# ===============================
def generate_report(df, tipe, zero_option):

    if tipe == "QS":
        df['PREMIUM'] = df['QS_CEDING']
        df['COMMISSION'] = df['KOMISI_QS']
        df['CLAIM'] = df['KLAIM_QS']
    else:
        df['PREMIUM'] = df['SP_CEDING']
        df['COMMISSION'] = df['KOMISI_SP']
        df['CLAIM'] = df['KLAIM_SP']
    
    df['AMOUNT'] = df['PREMIUM'] - df['COMMISSION']

    grouped = (
        df.groupby(['CURRENCY','COB','UY'])
        .sum(numeric_only=True)
        .reset_index()
        .sort_values(['CURRENCY','COB','UY'])
    )

    rows = []

    # ============================
    # LOOP CURRENCY
    # ============================
    for curr, df_curr in grouped.groupby('CURRENCY'):

        # 🔥 FILTER ZERO (LEVEL CURRENCY)
        if zero_option == "Hide Zero Rows":
            df_curr = df_curr[
                ~(
                    (df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']] == 0)
                    .all(axis=1)
                )
            ]

        if df_curr.empty:
            continue

        first_row = True

        # ============================
        # LOOP COB
        # ============================
        for cob, df_cob in df_curr.groupby('COB'):

            # 🔥 FILTER ZERO (LEVEL COB)
            if zero_option == "Hide Zero Rows":
                df_cob = df_cob[
                    ~(
                        (df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']] == 0)
                        .all(axis=1)
                    )
                ]

            if df_cob.empty:
                continue

            # ============================
            # DETAIL ROW
            # ============================
            for _, r in df_cob.iterrows():
                rows.append([
                    curr if first_row else "",
                    cob,
                    r['UY'],
                    r['PREMIUM'],
                    r['COMMISSION'],
                    r['CLAIM'],
                    r['AMOUNT']
                ])
                first_row = False

            # ============================
            # SUBTOTAL COB
            # ============================
            subtotal = df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
            rows.append(["", f"{cob} TOTAL", "", *subtotal])

        # ============================
        # TOTAL CURRENCY
        # ============================
        total_curr = df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
        rows.append([f"{curr} TOTAL","","", *total_curr])

        # SPASI
        rows.append(["","","","","","",""])

    return pd.DataFrame(
        rows,
        columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','AMOUNT']
    )
report_qs = generate_report(df.copy(), "QS", zero_option)
report_sp = generate_report(df.copy(), "SP", zero_option)

st.dataframe(report_qs)

# ===============================
# INPUT
# ===============================
ref_qs = st.text_input("Ref No QS")
ref_sp = st.text_input("Ref No SL")
note = st.text_area("Note")
file_name = st.text_input("Nama file", value="SOA_Report")

# ===============================
# EXPORT
# ===============================
output = io.BytesIO()

from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image

header_fill = PatternFill("solid", fgColor="000000")
grey_fill = PatternFill("solid", fgColor="D9D9D9")
white_fill = PatternFill("solid", fgColor="FFFFFF")
no_border = Border()

def write_sheet(writer, data, name, tipe, ref):

    data.to_excel(writer, index=False, sheet_name=name, startrow=12)
    ws = writer.sheets[name]

        # ===============================
    # LOGO (optional)
    # ===============================
    try:
        logo = Image("askrindo.jpg")
        logo.height = 60
        logo.width = 140
        ws.add_image(logo, "A1")
    except:
        pass

    # ===============================
    # TITLE
    # ===============================
    ws.merge_cells('A4:G4')
    ws['A4'] = "STATEMENT OF ACCOUNT"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A5:G5')
    ws['A5'] = f"Ref No. {ref}"
    ws['A5'].font = Font(bold=True)
    ws['A5'].alignment = Alignment(horizontal='center')

    # ===============================
    # HEADER INFO
    # ===============================
    ws['A7'] = "Treaty Year  :"; ws['B7'] = year
    ws['A8'] = "Quarter      :"; ws['B8'] = f"{quarter} {tipe}"
    ws['A9'] = "For Months   :"; ws['B9'] = months_text
    ws['A10'] = "Broker       :"; ws['B10'] = selected_broker
    
    # HEADER TABLE
    for col in "ABCDEFG":
        cell = ws[f"{col}13"]
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = no_border

    current_currency = None

    for row in range(14, ws.max_row+1):

        val_curr = ws[f"A{row}"].value
        val_cob  = ws[f"B{row}"].value

        # DEFAULT PUTIH
        for col in "ABCDEFG":
            ws[f"{col}{row}"].fill = white_fill
            ws[f"{col}{row}"].border = no_border

        if all(ws[f"{col}{row}"].value in ["", None] for col in "ABCDEFG"):
            current_currency = None
            continue

        # ALIGNMENT
        ws[f"A{row}"].alignment = Alignment(horizontal='left')
        ws[f"B{row}"].alignment = Alignment(horizontal='left')
        ws[f"C{row}"].alignment = Alignment(horizontal='center')

        # DETECT CURRENCY
        if val_curr not in ["", None] and "TOTAL" not in str(val_curr):
            current_currency = val_curr

        # KOLOM A GREY
        if current_currency and val_cob not in ["", None]:
            ws[f"A{row}"].fill = grey_fill

        # TOTAL CURRENCY FULL GREY
        if val_curr and "TOTAL" in str(val_curr):
            for col in "ABCDEFG":
                ws[f"{col}{row}"].fill = grey_fill
            current_currency = None

        # BOLD
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)

        if "TOTAL" in str(val_cob) or "TOTAL" in str(val_curr):
            for col in "ABCDEFG":
                ws[f"{col}{row}"].font = Font(bold=True)

        # FORMAT ANGKA
        for col in ['D','E','F','G']:
            ws[f"{col}{row}"].number_format = '#,##0.00;[Red](#,##0.00)'

    # NOTE
    last = ws.max_row + 2
    ws[f"A{last}"] = "Note :"
    ws[f"B{last}"] = note

with pd.ExcelWriter(output, engine='openpyxl') as writer:
    write_sheet(writer, report_qs, "QS Report", "Quota Share", ref_qs)
    write_sheet(writer, report_sp, "SL Report", "Surplus", ref_sp)

st.download_button(
    "⬇️ Download Report",
    data=output.getvalue(),
    file_name=f"{file_name}.xlsx"
)

# import streamlit as st
# import pandas as pd
# import io

# st.set_page_config(page_title="SOA Report", layout="wide")
# st.title("📑 SOA Report Generator")

# # ===============================
# # UPLOAD
# # ===============================
# file = st.file_uploader("Upload File SOA", type=["xlsx"])
# if not file:
#     st.stop()

# excel_file = pd.ExcelFile(file)
# sheet = st.selectbox("Pilih Sheet", excel_file.sheet_names)
# df = pd.read_excel(excel_file, sheet_name=sheet)

# # ===============================
# # NORMALISASI KOLOM
# # ===============================
# df.columns = df.columns.str.strip().str.lower()

# mapping = {
#     'prod':'PROD','cob':'COB','uy':'UY','curr':'CURRENCY',
#     'broker':'BROKER','qs_ceding':'QS_CEDING','sp_ceding':'SP_CEDING',
#     'komisi_qs':'KOMISI_QS','komisi_sp':'KOMISI_SP'
# }
# df = df.rename(columns=mapping)

# # ===============================
# # FIX TEXT (PENTING BANGET)
# # ===============================
# for col in ['CURRENCY', 'COB', 'BROKER']:
#     if col in df.columns:
#         df[col] = (
#             df[col]
#             .astype(str)
#             .str.strip()        # hapus spasi depan belakang
#             .str.upper()        # samakan huruf
#         )

# df = df[df['CURRENCY'] != ""]
# df = df[df['CURRENCY'].notna()]

# # ===============================
# # FILTER BROKER
# # ===============================
# broker_list = df['BROKER'].dropna().unique().tolist()
# broker_list.insert(0, "ALL")
# selected_broker = st.selectbox("Pilih Broker", broker_list)

# if selected_broker != "ALL":
#     df = df[df['BROKER'] == selected_broker]

# # ===============================
# # FILTER LT
# # ===============================
# lt_option = st.selectbox("Filter Long Term", ["ALL", "LT", "NON-LT"])

# if lt_option != "ALL":
#     if lt_option == "LT":
#         df = df[df['COB'].str.contains("LT", case=False, na=False)]
#     else:
#         df = df[~df['COB'].str.contains("LT", case=False, na=False)]

# # ===============================
# # FILTER ZERO ROW
# # ===============================
# zero_option = st.selectbox(
#     "Tampilkan Baris Nol",
#     ["Show All", "Hide Zero Rows"]
# )

# # ===============================
# # CLEAN NUMERIC
# # ===============================
# num_cols = ['QS_CEDING','SP_CEDING','KOMISI_QS','KOMISI_SP']

# for col in num_cols:
#     df[col] = (
#         df[col].astype(str)
#         .str.replace('.', '', regex=False)
#         .str.replace(',', '.', regex=False)
#     )
#     df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# # ===============================
# # PARSE PROD
# # ===============================
# def parse_prod(x):
#     try:
#         x = str(x)
#         return int(x[:4]), int(x[-2:])
#     except:
#         return None, None

# df[['YEAR','MONTH']] = df['PROD'].apply(lambda x: pd.Series(parse_prod(x)))
# df = df.dropna(subset=['YEAR','MONTH'])

# # ===============================
# # MONTH + QUARTER
# # ===============================
# month_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
#              7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# min_m = int(df['MONTH'].min())
# max_m = int(df['MONTH'].max())
# year = int(df['YEAR'].mode()[0])

# months_text = f"{month_map[min_m]} - {month_map[max_m]} {year}"

# def get_quarter(m):
#     if m <= 3: return "I"
#     elif m <= 6: return "II"
#     elif m <= 9: return "III"
#     else: return "IV"

# quarter = get_quarter(max_m)

# # ===============================
# # GENERATE REPORT
# # ===============================
# def generate_report(df, tipe, zero_option):

#     if tipe == "QS":
#         df['PREMIUM'] = df['QS_CEDING']
#         df['COMMISSION'] = df['KOMISI_QS']
#     else:
#         df['PREMIUM'] = df['SP_CEDING']
#         df['COMMISSION'] = df['KOMISI_SP']

#     df['CLAIM'] = 0
#     df['AMOUNT'] = df['PREMIUM'] - df['COMMISSION']

#     grouped = df.groupby(['CURRENCY','COB','UY']).sum(numeric_only=True).reset_index()

#     rows = []

#     for curr in grouped['CURRENCY'].unique():
#         df_curr = grouped[grouped['CURRENCY'] == curr]

#         first_row = True

#         for cob in df_curr['COB'].unique():
#             df_cob = df_curr[df_curr['COB'] == cob]

#             for _, r in df_cob.iterrows():

#                 # 🔥 FILTER ZERO ROW
#                 if zero_option == "Hide Zero Rows":
#                     if (
#                         r['PREMIUM'] == 0 and
#                         r['COMMISSION'] == 0 and
#                         r['CLAIM'] == 0 and
#                         r['AMOUNT'] == 0
#                     ):
#                         continue  # skip baris ini
#             for _, r in df_cob.iterrows():
#                 rows.append([
#                     curr if first_row else "",
#                     cob,
#                     r['UY'],
#                     r['PREMIUM'],
#                     r['COMMISSION'],
#                     r['CLAIM'],
#                     r['AMOUNT']
#                 ])
#                 first_row = False
                

#             subtotal = df_cob[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
#             rows.append(["", f"{cob} TOTAL", "", *subtotal])

#         total_curr = df_curr[['PREMIUM','COMMISSION','CLAIM','AMOUNT']].sum()
#         rows.append([f"{curr} TOTAL","","", *total_curr])
#         rows.append(["","","","","","",""])

#     return pd.DataFrame(rows,
#         columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','AMOUNT'])

# report_qs = generate_report(df.copy(), "QS")
# report_sp = generate_report(df.copy(), "SP")

# st.dataframe(report_qs)

# # ===============================
# # INPUT
# # ===============================
# ref_qs = st.text_input("Ref No QS")
# ref_sp = st.text_input("Ref No SL")
# note = st.text_area("Note")
# file_name = st.text_input("Nama file", value="SOA_Report")

# # ===============================
# # EXPORT
# # ===============================
# output = io.BytesIO()

# from openpyxl.styles import Font, Alignment, PatternFill, Border
# from openpyxl.drawing.image import Image

# header_fill = PatternFill("solid", fgColor="000000")
# grey_fill = PatternFill("solid", fgColor="D9D9D9")
# no_border = Border()

# def write_sheet(writer, data, name, tipe, ref):

#     data.to_excel(writer, index=False, sheet_name=name, startrow=12)
#     ws = writer.sheets[name]

#     # LOGO
#     try:
#         logo = Image("askrindo.jpg")
#         logo.height = 60
#         logo.width = 140
#         ws.add_image(logo, "A1")
#     except:
#         pass

#     # TITLE
#     ws.merge_cells('A4:G4')
#     ws['A4'] = "STATEMENT OF ACCOUNT"
#     ws['A4'].font = Font(bold=True, size=14)
#     ws['A4'].alignment = Alignment(horizontal='center')

#     ws.merge_cells('A5:G5')
#     ws['A5'] = f"Ref No. {ref}"
#     ws['A5'].font = Font(bold=True)
#     ws['A5'].alignment = Alignment(horizontal='center')

#     # HEADER INFO
#     ws['A7'] = "Treaty Year  :"; ws['B7'] = year
#     ws['A8'] = "Quarter      :"; ws['B8'] = f"{quarter} {tipe}"
#     ws['A9'] = "For Months   :"; ws['B9'] = months_text
#     ws['A10'] = "Broker       :"; ws['B10'] = selected_broker

#     # HEADER TABLE
#     for col in "ABCDEFG":
#         cell = ws[f"{col}13"]
#         cell.fill = header_fill
#         cell.font = Font(color="FFFFFF", bold=True)
#         cell.border = no_border

#     current_currency = None

#     # BODY STYLE
#     from openpyxl.styles import PatternFill

#     white_fill = PatternFill("solid", fgColor="FFFFFF")

#     current_currency = None
    
#     for row in range(14, ws.max_row+1):
    
#         val_curr = ws[f"A{row}"].value
#         val_cob  = ws[f"B{row}"].value
    
#         # ============================
#         # DEFAULT PUTIH TANPA GARIS
#         # ============================
#         for col in "ABCDEFG":
#             cell = ws[f"{col}{row}"]
#             cell.fill = white_fill
#             cell.border = no_border
    
#         # ============================
#         # SKIP BARIS KOSONG
#         # ============================
#         if all(ws[f"{col}{row}"].value in ["", None] for col in "ABCDEFG"):
#             current_currency = None
#             continue
    
#         # ============================
#         # ALIGNMENT
#         # ============================
#         ws[f"A{row}"].alignment = Alignment(horizontal='left')
#         ws[f"B{row}"].alignment = Alignment(horizontal='left')
#         ws[f"C{row}"].alignment = Alignment(horizontal='center')
    
#         # ============================
#         # DETECT CURRENCY
#         # ============================
#         if val_curr not in ["", None] and "TOTAL" not in str(val_curr):
#             current_currency = val_curr
    
#         # ============================
#         # KOLOM A (VERTICAL GREY)
#         # ============================
#         if current_currency and val_cob not in ["", None]:
#             ws[f"A{row}"].fill = grey_fill
    
#         # ============================
#         # TOTAL CURRENCY (FULL GREY)
#         # ============================
#         if val_curr and "TOTAL" in str(val_curr):
#             for col in "ABCDEFG":
#                 ws[f"{col}{row}"].fill = grey_fill
#             current_currency = None
    
#         # ============================
#         # BOLD
#         # ============================
#         ws[f"A{row}"].font = Font(bold=True)
#         ws[f"B{row}"].font = Font(bold=True)
    
#         if "TOTAL" in str(val_cob) or "TOTAL" in str(val_curr):
#             for col in "ABCDEFG":
#                 ws[f"{col}{row}"].font = Font(bold=True)
    
#         # ============================
#         # FORMAT ANGKA
#         # ============================
#         for col in ['D','E','F','G']:
#             ws[f"{col}{row}"].number_format = '#,##0.00;[Red](#,##0.00)'
            
#         # for row in range(14, ws.max_row+1):
    
#         #     val_curr = ws[f"A{row}"].value
#         #     val_cob  = ws[f"B{row}"].value
    
#         #     # alignment
#         #     ws[f"A{row}"].alignment = Alignment(horizontal='left')
#         #     ws[f"B{row}"].alignment = Alignment(horizontal='left')
#         #     ws[f"C{row}"].alignment = Alignment(horizontal='center')
    
#         #     # detect currency block
#         #     if val_curr not in ["", None] and "TOTAL" not in str(val_curr):
#         #         current_currency = val_curr
    
#         #     # kolom A grey (vertical)
#         #     if current_currency:
#         #         ws[f"A{row}"].fill = grey_fill
    
#         #     # total currency → full grey
#         #     if val_curr and "TOTAL" in str(val_curr):
#         #         for col in "ABCDEFG":
#         #             ws[f"{col}{row}"].fill = grey_fill
#         #         current_currency = None
    
#         #     # bold COB + total
#         #     ws[f"B{row}"].font = Font(bold=True)
    
#         #     if "TOTAL" in str(val_cob) or "TOTAL" in str(val_curr):
#         #         for col in "ABCDEFG":
#         #             ws[f"{col}{row}"].font = Font(bold=True)
    
#         #     # number format
#         #     for col in ['D','E','F','G']:
#         #         ws[f"{col}{row}"].number_format = '#,##0.00;[Red](#,##0.00)'
    
#             # remove border
#             for col in "ABCDEFG":
#                 ws[f"{col}{row}"].border = no_border

#     # NOTE
#     last = ws.max_row + 2
#     ws[f"A{last}"] = "Note :"
#     ws[f"A{last}"].font = Font(bold=True)
#     ws[f"B{last}"] = note
#     ws[f"B{last}"].font = Font(bold=True)

# with pd.ExcelWriter(output, engine='openpyxl') as writer:
#     write_sheet(writer, report_qs, "QS Report", "Quota Share", ref_qs)
#     write_sheet(writer, report_sp, "SL Report", "Surplus", ref_sp)

# st.download_button(
#     "⬇️ Download Report",
#     data=output.getvalue(),
#     file_name=f"{file_name}.xlsx"
# )


