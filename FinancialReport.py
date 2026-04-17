import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image

st.set_page_config(page_title="SOA Report Generator", layout="wide")

# ===============================
# HEADER UI (STYLE BARU)
# ===============================
col1, col2 = st.columns([1, 6])

with col1:
    st.image("askrindo.jpg", width=120)

with col2:
    st.title("📑 SOA Financial Report Generator")

# ===============================
# UPLOAD
# ===============================
file = st.file_uploader("Upload File SOA", type=["xlsx","xlsb"])
if not file:
    st.stop()

# detect file
if file.name.endswith(".xlsb"):
    excel_file = pd.ExcelFile(file, engine="pyxlsb")
else:
    excel_file = pd.ExcelFile(file)

sheet = st.selectbox("Pilih Sheet", excel_file.sheet_names)

if file.name.endswith(".xlsb"):
    df = pd.read_excel(file, sheet_name=sheet, engine="pyxlsb")
else:
    df = pd.read_excel(file, sheet_name=sheet)

def auto_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[col_letter].width = adjusted_width
# ===============================
# NORMALISASI (ASLI)
# ===============================
df.columns = df.columns.str.strip().str.lower()

mapping = {
    'prod':'PROD','product':'COB','cob':'COB','uy':'UY',
    'valuta':'CURRENCY','curr':'CURRENCY','currency':'CURRENCY',
    'broker':'BROKER',
    'premi_panel_qs':'PREMI_PANEL_QS',
    'premi_panel_sp':'PREMI_PANEL_SP',
    'komisi_panel_qs':'KOMISI_PANEL_QS',
    'komisi_panel_sp':'KOMISI_PANEL_SP',
    'klaim_panel_qs':'KLAIM_PANEL_QS',
    'klaim_panel_sp':'KLAIM_PANEL_SP',
    'recoveries_panel_qs':'RECOVERY_PANEL_QS',
    'recoveries_panel_sp':'RECOVERY_PANEL_SP'
}
df = df.rename(columns=mapping)

# ===============================
# FILTER UI (STYLE BARU)
# ===============================
st.subheader("Pilih Broker")
broker_list = sorted(df["BROKER"].dropna().unique().tolist())

selected_broker = broker_list if st.checkbox("ALL BROKER", True) else [
    b for b in broker_list if st.checkbox(b)
]
df = df[df["BROKER"].isin(selected_broker)]

st.subheader("Pilih COB")
cob_list = sorted(df["COB"].dropna().unique().tolist())

selected_cob = cob_list if st.checkbox("ALL COB", True) else [
    c for c in cob_list if st.checkbox(c)
]
df = df[df["COB"].isin(selected_cob)]

st.subheader("Pilih UW Year")
uy_list = sorted(df["UY"].dropna().unique().tolist())

selected_uy = uy_list if st.checkbox("ALL UW YEAR", True) else [
    u for u in uy_list if st.checkbox(str(u))
]
df = df[df["UY"].isin(selected_uy)]

zero_option = st.selectbox("Tampilkan Baris Nol", ["Show All", "Hide Zero Rows"])

# ===============================
# DATE
# ===============================
df['MONTH'] = pd.to_numeric(df['bulan'], errors='coerce')
df['YEAR']  = pd.to_numeric(df['tahun'], errors='coerce')
df = df.dropna(subset=['MONTH','YEAR'])

month_map = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
             7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

min_m = int(df['MONTH'].min())
max_m = int(df['MONTH'].max())
year = int(df['YEAR'].mode()[0])

months_text = f"{month_map[min_m]} - {month_map[max_m]} {year}"

def format_number(ws, start_row, end_row):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=4, max_col=8):
        for cell in row:
            if cell.value is not None:
                # ✅ format Indonesia
                cell.number_format = '#,##0.00;[Red](#,##0.00)'

# ===============================
# GENERATE REPORT (ASLI KAMU)
# ===============================
def generate_report(df, tipe, zero_option):

    if tipe == "QS":
        df['PREMIUM'] = df['PREMI_PANEL_QS']
        df['COMMISSION'] = df['KOMISI_PANEL_QS']
        df['CLAIM'] = df['KLAIM_PANEL_QS']
        df['RECOVERY'] = df['RECOVERY_PANEL_QS']
    else:
        df['PREMIUM'] = df['PREMI_PANEL_SP']
        df['COMMISSION'] = df['KOMISI_PANEL_SP']
        df['CLAIM'] = df['KLAIM_PANEL_SP']
        df['RECOVERY'] = df['RECOVERY_PANEL_SP']

    df['AMOUNT'] = df['PREMIUM'] - df['COMMISSION'] - df['CLAIM'] + df['RECOVERY']

    grouped = (
        df.groupby(['CURRENCY','COB','UY'])
        .sum(numeric_only=True)
        .reset_index()
        .sort_values(['CURRENCY','COB','UY'])
    )

    rows = []

    for curr, df_curr in grouped.groupby('CURRENCY'):

        first_row = True

        for cob, df_cob in df_curr.groupby('COB'):
            first_cob = True

            for _, r in df_cob.iterrows():
                rows.append([
                    curr if first_row else "",
                    cob if first_cob else "",
                    r['UY'],
                    r['PREMIUM'],
                    r['COMMISSION'],
                    r['CLAIM'],
                    r['RECOVERY'],
                    r['AMOUNT']
                ])
                first_row = False
                first_cob = False

            subtotal = df_cob[['PREMIUM','COMMISSION','CLAIM','RECOVERY','AMOUNT']].sum()
            rows.append(["", f"{cob} TOTAL", "", *subtotal])

        total = df_curr[['PREMIUM','COMMISSION','CLAIM','RECOVERY','AMOUNT']].sum()
        rows.append([f"{curr} TOTAL","","", *total])
        rows.append(["","","","","","","",""])

    return pd.DataFrame(
        rows,
        columns=['CURRENCY','COB','UW YEAR','PREMIUM','COMMISSION','CLAIM','RECOVERY','AMOUNT']
    )
import re

def parse_ref(ref):
    match = re.match(r"(\d+)(/.+)", ref)
    if match:
        return int(match.group(1)), match.group(2)
    return 1, ""

def build_ref(num, suffix):
    return f"{num:02d}{suffix}"

# ===============================
# INPUT USER
# ===============================
st.subheader("Header Info")

treaty_year_input = st.text_input("Treaty Year", value=str(year))
quarter_input = st.selectbox("Quarter", ["I","II","III","IV"])
remarks_input = st.text_input("Remarks", value="-")

ref_input = st.text_input("Ref No Awal (contoh: 83/UDWR/III/2025)")
ttd_date = st.text_input("Tanggal TTD (contoh: Jakarta, 15 Januari 2026)")
ttd_name = st.text_input("Nama Penandatangan")
ttd_jabatan = st.text_input("Jabatan Penandatangan")
file_name_input = st.text_input("Nama File Output", value="SOA_REPORT")

def add_signature(ws, start_row, broker, date_text, nama, jabatan):

    # ======================
    # KIRI
    # ======================
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    ws.cell(row=start_row, column=1).value = "Agreed and Approved by Reinsurer"

    # LANGSUNG DI BAWAH (tanpa lompat)
    ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=3)
    cell_broker = ws.cell(row=start_row+1, column=1)
    cell_broker.value = broker
    cell_broker.font = Font(bold=True)

    # ======================
    # KANAN
    # ======================
    for r in range(start_row, start_row+6):
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)

    # tanggal
    ws.cell(row=start_row, column=5).value = date_text
    ws.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')

    # company
    ws.cell(row=start_row+1, column=5).value = "PT. Asuransi Kredit Indonesia"
    ws.cell(row=start_row+1, column=5).font = Font(bold=True)
    ws.cell(row=start_row+1, column=5).alignment = Alignment(horizontal='center')

    # divisi
    ws.cell(row=start_row+2, column=5).value = "Underwriting & Reinsurance Division"
    ws.cell(row=start_row+2, column=5).alignment = Alignment(horizontal='center')

        # nama (turun 2 baris)
    ws.cell(row=start_row+6, column=5).value = nama
    ws.cell(row=start_row+6, column=5).font = Font(underline="single")
    ws.cell(row=start_row+6, column=5).alignment = Alignment(horizontal='center')
    
    # jabatan (turun 2 baris)
    ws.cell(row=start_row+7, column=5).value = jabatan
    ws.cell(row=start_row+7, column=5).alignment = Alignment(horizontal='center')
    
grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

def style_total_rows(ws, start_row, end_row):
    for r in range(start_row, end_row):
        val_a = str(ws.cell(r,1).value)

        # ✅ HANYA currency total (kolom A)
        if "TOTAL" in val_a:
            for c in range(1,9):
                cell = ws.cell(r,c)
                cell.fill = grey_fill
                cell.font = Font(bold=True)

        # ❌ COB TOTAL → jangan bold

from openpyxl.styles import Side

thin = Side(style='thin')

def add_total_borders(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):

        val_a = str(ws.cell(r, 1).value)
        val_b = str(ws.cell(r, 2).value)

        # =========================
        # ✅ CURRENCY TOTAL (kolom A)
        # =========================
        if "TOTAL" in val_a:
            for c in range(1, 9):
                cell = ws.cell(r, c)
                cell.border = Border(
                    top=thin,
                    bottom=thin
                )

        # =========================
        # ✅ COB TOTAL (kolom B)
        # =========================
        elif "TOTAL" in val_b:
            for c in range(2, 9):  # mulai dari kolom B
                cell = ws.cell(r, c)
                cell.border = Border(
                    top=thin,
                    bottom=thin
                )
# ===============================
# EXPORT EXCEL ONLY
# ===============================
if st.button("⬇️ Generate Excel"):

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # 🔥 TARUH DI SINI (SEBELUM LOOP)
        start_num, suffix = parse_ref(ref_input)
        ref_counter = start_num
    
        for broker in selected_broker:
    
            df_b = df[df["BROKER"] == broker]
    
            qs = generate_report(df_b.copy(), "QS", zero_option)
            sp = generate_report(df_b.copy(), "SP", zero_option)
    
            sheet_name = str(broker)[:31]
    
            # ======================
            # QS REF
            # ======================
            ref_qs = build_ref(ref_counter, suffix)
            ref_counter += 1
    
            header_info_start = 7
            remarks_row = header_info_start + 4   # karena 5 header (0–4)
            table_start_qs = remarks_row + 1      # 1 spasi setelah remarks
            
            qs.to_excel(writer, sheet_name=sheet_name, startrow=table_start_qs, index=False)           
            ws = writer.sheets[sheet_name]
            style_total_rows(ws, 14, 14 + len(qs))
            # ===============================
            # HEADER TABLE QS (HITAM)
            # ===============================
            header_fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
            
            header_row_qs = table_start_qs + 1 # karena startrow=12 → header di baris 13
            
            for col in range(1, 9):
                cell = ws.cell(row=header_row_qs, column=col)
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            add_total_borders(ws, header_row_qs + 1, header_row_qs + len(qs))
            format_number(ws, header_row_qs + 1, header_row_qs + len(qs))
            auto_column_width(ws)

            # ===============================
            # TITLE
            # ===============================
            ws.merge_cells('A4:H4')
            ws['A4'] = "STATEMENT OF ACCOUNT"
            ws['A4'].font = Font(bold=True, size=14)
            ws['A4'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A5:H5')
            ws['A5'] = f"Ref No. {ref_qs}"
            ws['A5'].alignment = Alignment(horizontal='center')
            
            # ===============================
            # HEADER INFO (RAPI SEJAJAR)
            # ===============================
            contract_name_qs = "Quota Share"
            
            headers = [
                ("Treaty Year", treaty_year_input),
                ("Quarter", quarter_input),
                ("Contract Name", contract_name_qs),
                ("For Months", months_text),
                ("Remarks", remarks_input),
            ]
            
            start_row = 7
            
            for i, (label, value) in enumerate(headers):
                r = start_row + i
            
                ws[f"A{r}"] = label
                ws[f"B{r}"] = f": {value}"
            
                ws[f"A{r}"].alignment = Alignment(horizontal="left")
                ws[f"B{r}"].alignment = Alignment(horizontal="left")
            # ===============================
            # TTD QS
            # ===============================
            # table_start_qs = 12
            table_end_qs = table_start_qs + len(qs) + 1
            
            end_qs = table_end_qs + 1   # 1 baris kosong

            add_signature(ws, end_qs, broker, ttd_date, ttd_name, ttd_jabatan)

            # # ===============================
            # # TITLE
            # # ===============================
            # ws.merge_cells('A4:H4')
            # ws['A4'] = "STATEMENT OF ACCOUNT"
            # ws['A4'].font = Font(bold=True, size=14)
            # ws['A4'].alignment = Alignment(horizontal='center')
            
            # ws.merge_cells('A5:H5')
            # ws['A5'] = f"Ref No. {ref_qs}"
            # ws['A5'].alignment = Alignment(horizontal='center')
            
            # ===============================
            # HEADER INFO (RAPI SEJAJAR)
            # ===============================
            contract_name_qs = "Quota Share"
            
            headers = [
                ("Treaty Year", treaty_year_input),
                ("Quarter", quarter_input),
                ("Contract Name", contract_name_qs),
                ("For Months", months_text),
                ("Remarks", remarks_input),
            ]
            
            start_row = 7
            
            for i, (label, value) in enumerate(headers):
                r = start_row + i
            
                ws[f"A{r}"] = label
                ws[f"B{r}"] = f": {value}"
            
                ws[f"A{r}"].alignment = Alignment(horizontal="left")
                ws[f"B{r}"].alignment = Alignment(horizontal="left")

            # ======================
            # SP REF
            # ======================
            ref_sp = build_ref(ref_counter, suffix)
            ref_counter += 1
            
            sp_start = end_qs + 6   # lebih rapih & konsisten
            # header info SP
            start_row_sp = sp_start + 3
            remarks_row_sp = start_row_sp + 4
            
            table_start_sp = remarks_row_sp + 1
            
            sp.to_excel(writer, sheet_name=sheet_name, startrow=table_start_sp, index=False)
            # 🔥 HITUNG POSISI AKHIR TABEL (INI YANG KAMU TANYA)
            table_end_sp = table_start_sp + len(sp) + 1
            end_sp = table_end_sp + 1
            from openpyxl.worksheet.pagebreak import Break  # taruh di atas file (import)
            header_row_sp = table_start_sp + 1
            style_total_rows(ws, header_row_sp + 1, header_row_sp + 1 + len(sp))
            # ===============================
            # HEADER TABLE SP (HITAM)
            # ===============================
     # karena startrow=sp_start+8
            
            for col in range(1, 9):
                cell = ws.cell(row=header_row_sp, column=col)
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            add_total_borders(ws, header_row_sp + 1, header_row_sp + len(sp))
            format_number(ws, header_row_sp + 1, header_row_sp + len(sp))
            auto_column_width(ws)
            # ===============================
            # SET A4 + PAGE BREAK
            # ===============================
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            # ===============================
            # PRINT SETTING A4 FIX
            # ===============================
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False
            
            # repeat header tiap page (biar rapi kalau panjang)
            ws.print_title_rows = f"{header_row_qs}:{header_row_qs}"
            
            # set print area full (biar gak kepotong)
            ws.print_area = f"A1:H{end_sp+5}"
            
            # page break sebelum SP dimulai
            from openpyxl.worksheet.pagebreak import Break
            
            ws.row_breaks = []  # reset dulu biar bersih
            ws.row_breaks.append(Break(id=sp_start))
            # ===============================
            # TITLE SP (SAMA KAYAK QS)
            # ===============================
            ws.merge_cells(f'A{sp_start}:H{sp_start}')
            ws[f'A{sp_start}'] = "STATEMENT OF ACCOUNT"
            ws[f'A{sp_start}'].font = Font(bold=True, size=14)
            ws[f'A{sp_start}'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells(f'A{sp_start+1}:H{sp_start+1}')
            ws[f'A{sp_start+1}'] = f"Ref No. {ref_sp}"
            ws[f'A{sp_start+1}'].alignment = Alignment(horizontal='center')
            
            # ===============================
            # HEADER INFO SP
            # ===============================
            contract_name_sp = "Surplus"
            
            headers_sp = [
                ("Treaty Year", treaty_year_input),
                ("Quarter", quarter_input),
                ("Contract Name", contract_name_sp),
                ("For Months", months_text),
                ("Remarks", remarks_input),
            ]
            
            start_row_sp = sp_start + 3
            
            for i, (label, value) in enumerate(headers_sp):
                r = start_row_sp + i
            
                ws[f"A{r}"] = label
                ws[f"B{r}"] = f": {value}"
            
                ws[f"A{r}"].alignment = Alignment(horizontal="left")
                ws[f"B{r}"].alignment = Alignment(horizontal="left")

            # ===============================
            # TTD SP
            # ===============================
            add_signature(ws, end_sp, broker, ttd_date, ttd_name, ttd_jabatan)

    st.download_button(
        "📥 Download Excel",
        data=output.getvalue(),
        file_name=f"{file_name_input}.xlsx"
    )
